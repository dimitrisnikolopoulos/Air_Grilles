#! /usr/bin/env python
# -*- coding: UTF-8 -*-
"""
Created on 05/06/2021 17:50

@author: Nikolopoulos Dimitris aka Entity
@email: nikoldimitris@gmail.com
"""

import os

import PySimpleGUI as sg
import numpy as np
import pandas as pd

import jinja2

import openpyxl
from openpyxl import load_workbook
from openpyxl.chart import LineChart, SurfaceChart, Reference
from openpyxl.chart.trendline import Trendline


def read_air_tunnel(filename):
    """ * """
    print('Inside Read Air Tunnel', filename)
    try:
        df = pd.read_csv(filename, engine='python', index_col='TimeString', usecols=['TimeString', 'VarName', 'VarValue'],
                         parse_dates=['TimeString'], infer_datetime_format=True, dayfirst=False, dtype={'VarName': np.object0})
        df = df[df['VarName'] != '$RT_OFF$']
        df = df[df['VarName'] != '$RT_COUNT$']
        data = df.pivot_table(index=df.index, columns=df['VarName'], aggfunc=np.sum)
        data.index = pd.to_datetime(data.index, errors='ignore')
        data.sort_index(inplace=True)

        if len(data.columns) == 9:
            data.rename(columns={'1': 'X (mm)', '2': 'Y (mm)', '3': 'Z (mm)', '4': 'In Air Speed (m/s)', '5': 'In Air Flow (m³/h)',
                                 '6': 'In Air Temp (°C)', '7': 'Out Air Speed (m/s)', '8': 'Out Dif Pressure (Pa)', '9': 'Fan (Hz)'},
                        inplace=True)
            data.columns = data.columns.droplevel(0)
            data.columns.name = None
            # data = data[data.Fan!=0]
            data = data.round({'X (mm)': 1, 'Y (mm)': 1, 'Z (mm)': 1})
        else:
            sg.Popup('Select Correct CSV File for AIR TUNNEL')
            data = pd.DataFrame()

        return data
    except ValueError:
        sg.Popup('Select Correct File')
        return pd.DataFrame()


def save_air_tunnel(df, filename):
    """ * """
    print('Inside Air Tunnel Save', filename)
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        workbook = writer.book
        df.to_excel(excel_writer=writer, sheet_name='Sheet1', float_format='%.4f', encoding='utf8')
        pivot = df.pivot_table(index=df['Fan (Hz)'], values=['In Air Flow (m³/h)', 'Out Air Speed (m/s)', 'Out Dif Pressure (Pa)'], aggfunc=np.mean)
        pivot['S (m²)'] = pivot['In Air Flow (m³/h)'] / (pivot['Out Air Speed (m/s)'] * 3600)
        pivot['<S> (m²)'] = pivot['S (m²)'].mean()
        pivot['dS (m²)'] = (pivot['S (m²)'] - pivot['<S> (m²)'])/pivot['<S> (m²)']
        pivot.to_excel(excel_writer=writer, sheet_name='Sheet2', float_format='%.4f', encoding='utf8')

        worksheet = writer.sheets['Sheet2']

        for _c in range(2, worksheet.max_row + 1):
            worksheet.cell(_c, 7).number_format = '0.00%'

        data1 = Reference(writer.sheets['Sheet2'], 3, 1, 3, worksheet.max_row)
        data2 = Reference(writer.sheets['Sheet2'], 4, 1, 4, worksheet.max_row)
        cats = Reference(writer.sheets['Sheet2'], 2, 2, 2, worksheet.max_row)

        c1 = LineChart()
        # c1.height = 10
        c1.title = 'Σχέση Παροχής και Ταχύτητας Εξόδου'
        c1.x_axis.title = 'Παροχή Αέρα Q (m³/h)'
        c1.y_axis.title = 'Ταχύτητα Εξόδου (m/s)'
        c1.add_data(data1, titles_from_data=True)
        c1.set_categories(cats)
        line1 = c1.series[0]
        line1.trendline = Trendline(dispEq=True, dispRSqr=True)
        line1.graphicalProperties.line = openpyxl.drawing.line.LineProperties(
            solidFill=openpyxl.drawing.colors.ColorChoice(prstClr='blue'), prstDash='dash')

        c2 = LineChart()
        # c2.height = 10
        c2.title = 'Πτώση Πίεσης Στομίου ΔP'
        c2.x_axis.title = 'Παροχή Αέρα Q (m³/h)'
        c2.y_axis.title = 'Πτώση Πίεσης ΔP (Pa)'
        c2.add_data(data2, titles_from_data=True)
        c2.set_categories(cats)
        line2 = c2.series[0]
        line2.trendline = Trendline(dispEq=True, dispRSqr=True)
        line2.graphicalProperties.line = openpyxl.drawing.line.LineProperties(
            solidFill=openpyxl.drawing.colors.ColorChoice(prstClr='red'), prstDash='dash')

        worksheet = workbook.create_sheet('Sheet3')
        worksheet.add_chart(c1, 'A1')

        worksheet = workbook.create_sheet('Sheet4')
        worksheet.add_chart(c2, 'A1')

        for freq in sorted(df['Fan (Hz)'].unique()):
            freq_data = df[['X (mm)', 'Y (mm)', 'Out Air Speed (m/s)']][df['Fan (Hz)'] == freq]
            freq_data = freq_data.pivot_table(index=freq_data['X (mm)'], columns=freq_data['Y (mm)'], aggfunc=np.mean)
            freq_data.index.name = 'X (mm) / Y (mm)'
            freq_data.columns = freq_data.columns.droplevel()
            freq_data.columns.name = None
            sheet_name = str(freq) + ' (Hz)'
            freq_data.to_excel(excel_writer=writer, sheet_name=sheet_name, float_format='%.4f', encoding='utf8')

            worksheet = writer.sheets[sheet_name]

            c_data = Reference(worksheet, 2, 1, worksheet.max_column, worksheet.max_row)
            c_label = Reference(worksheet, 1, 2, 1, worksheet.max_row)

            c = SurfaceChart()
            c.add_data(c_data, titles_from_data=True)
            c.set_categories(c_label)

            # c.title = 'WS 100 10*10 U (m/s)'
            c.title = 'Μέση Ταχύτητα Εξόδου (m/s)'
            c.x_axis.title = 'X Axis (mm)'
            c.y_axis.title = 'Y Axis (mm)'
            # c.style = 13
            c.height = 15
            c.width = 20

            worksheet.add_chart(c, 'K1')


def coloring(x):
    """ * """
    color = 'black'
    if x == 0.2:
        color = 'blue'
    elif x == 0.3:
        color = 'red'
    elif x == 0.5:
        color = 'green'

    return "color : {}".format(color)


def bold(x):
    """ * """
    if x in (0.2, 0.3, 0.5):
        return 'font-weight: bold'


def read_range_room(r_filename):
    """ * """
    print("Inside Read Range Room", r_filename)
    rdf = pd.read_csv(r_filename, engine='python',
                      index_col='TimeString', usecols=['TimeString', 'VarName', 'VarValue'],
                      parse_dates=True, infer_datetime_format=True, dayfirst=True,
                      dtype={'VarValue': np.float64},
                      converters={'VarName': lambda x: x[-2:] if '$' not in x else x}
                      )
    rdf = rdf[rdf['VarName'] != '$RT_OFF$']
    rdf = rdf[rdf['VarName'] != '$RT_COUNT$']

    rdata = rdf.pivot_table(values='VarValue', index='TimeString', columns='VarName')
    rdata.index = pd.to_datetime(rdata.index, errors='ignore')
    rdata.sort_index(inplace=True)

    if len(rdata.columns) == 10:
        # rdata.index.name = None
        rdata.columns.name = None
        rdata.rename(
            columns={'01': 'X (mm)', '02': 'Y (mm)', '03': 'Z (mm)', '04': 'Inlet Air Speed (m/s)', '05': 'Inlet Air Flow (m³/h)',
                     '06': 'Inlet Air Temp (°C)', '07': 'Outlet Air Speed (m/s)', '08': 'Outlet Air Temp (°C)', '09': 'Temp Diff (°C)',
                     '10': 'Fan (Hz)'}, inplace=True)
        rdata['Outlet Air Speed (m/s)'][(rdata['Outlet Air Speed (m/s)'] >= 0.19) & (rdata['Outlet Air Speed (m/s)'] <= 0.21)] = 0.2
        rdata['Outlet Air Speed (m/s)'][(rdata['Outlet Air Speed (m/s)'] >= 0.29) & (rdata['Outlet Air Speed (m/s)'] <= 0.31)] = 0.3
        rdata['Outlet Air Speed (m/s)'][(rdata['Outlet Air Speed (m/s)'] >= 0.49) & (rdata['Outlet Air Speed (m/s)'] <= 0.51)] = 0.5
        rdata['0.2'] = ''
        rdata['0.3'] = ''
        rdata['0.5'] = ''
    else:
        sg.Popup('Select Correct CSV File for RANGE ROOM')
        rdata = pd.DataFrame()

    return rdata


def save_range_room(rdf, r_filename):
    """ * """
    print('Inside Range Room Save', r_filename)
    with pd.ExcelWriter(r_filename, engine='openpyxl') as writer:
        rdf.style.applymap(coloring, subset='Outlet Air Speed (m/s)').applymap(bold, subset='Outlet Air Speed (m/s)').apply(
            lambda x: ['background-color: royalblue'] * x.shape[0] if '0.2' in x.name else
            ['background-color: orangered'] * x.shape[0] if '0.3' in x.name else
            ['background-color: greenyellow'] * x.shape[0] if '0.5' in x.name else
            ['background-color: white'] * x.shape[0], axis=0, subset=['0.2', '0.3', '0.5']).to_excel(
            excel_writer=writer, sheet_name='Sheet1', float_format='%.4f', encoding='utf8')


def process_file(p_filename):
    """ * """
    print('Inside Process', p_filename)
    try:
        df2 = pd.read_excel(p_filename, sheet_name='Sheet1', engine='openpyxl', index_col='Z (mm)',
                            usecols=['Z (mm)', 'Inlet Air Flow (m³/h)', 'Fan (Hz)', '0.2']).dropna().sort_values(
            by='Z (mm)', axis=0).drop(columns='0.2').rename(columns={'Inlet Air Flow (m³/h)': 'InAirFlow_0.2'})
        df3 = pd.read_excel(p_filename, sheet_name='Sheet1', engine='openpyxl', index_col='Z (mm)',
                            usecols=['Z (mm)', 'Inlet Air Flow (m³/h)', 'Fan (Hz)', '0.3']).dropna().sort_values(
            by='Z (mm)', axis=0).drop(columns='0.3').rename(columns={'Inlet Air Flow (m³/h)': 'InAirFlow_0.3'})
        df5 = pd.read_excel(p_filename, sheet_name='Sheet1', engine='openpyxl', index_col='Z (mm)',
                            usecols=['Z (mm)', 'Inlet Air Flow (m³/h)', 'Fan (Hz)', '0.5']).dropna().sort_values(
            by='Z (mm)', axis=0).drop(columns='0.5').rename(columns={'Inlet Air Flow (m³/h)': 'InAirFlow_0.5'})

        book = load_workbook(p_filename)
        with pd.ExcelWriter(p_filename, engine='openpyxl') as writer:
            writer.book = book
            if not df2.empty:
                # df2.index.name = '0.2'
                df2.to_excel(excel_writer=writer, sheet_name='Sheet2', startcol=0, float_format='%.4f', encoding='utf8')
            if not df3.empty:
                # df3.index.name = '0.3'
                df3.to_excel(excel_writer=writer, sheet_name='Sheet2', startcol=5, float_format='%.4f', encoding='utf8')
            if not df5.empty:
                # df5.index.name = '0.5'
                df5.to_excel(excel_writer=writer, sheet_name='Sheet2', startcol=10, float_format='%.4f', encoding='utf8')

            worksheet = writer.sheets['Sheet2']

            data1 = Reference(writer.sheets['Sheet2'], 2, 1, 2, worksheet.max_row) if worksheet.cell(2, 2).value else None
            data2 = Reference(writer.sheets['Sheet2'], 7, 1, 7, worksheet.max_row) if worksheet.cell(2, 7).value else None
            data3 = Reference(writer.sheets['Sheet2'], 12, 1, 12, worksheet.max_row) if worksheet.cell(2, 12).value else None

            cats = Reference(writer.sheets['Sheet2'], 1, 2, 1, worksheet.max_row)

            bl = LineChart()
            # bl.height = 10
            bl.title = 'Βεληνεκές T'
            bl.x_axis.title = 'Παροχή Αέρα Q (m³/h)'
            bl.y_axis.title = 'Βεληνεκές T (m)'
            if data1:
                bl.add_data(data1, titles_from_data=True)
                line_02 = bl.series[0]
                line_02.graphicalProperties.line = openpyxl.drawing.line.LineProperties(solidFill=openpyxl.drawing.colors.ColorChoice(prstClr='blue'))
                line_02.trendline = Trendline()
            if data2:
                bl.add_data(data2, titles_from_data=True)
                line_03 = bl.series[1]
                line_03.trendline = Trendline()
                line_03.graphicalProperties.line = openpyxl.drawing.line.LineProperties(solidFill=openpyxl.drawing.colors.ColorChoice(prstClr='red'))
            if data3:
                bl.add_data(data3, titles_from_data=True)
                line_05 = bl.series[2]
                line_05.graphicalProperties.line = openpyxl.drawing.line.LineProperties(solidFill=openpyxl.drawing.colors.ColorChoice(prstClr='green'))
                line_05.trendline = Trendline()

            bl.set_categories(cats)

            bl.y_axis.scaling.logBase = 10

            worksheet.add_chart(bl, 'F10')

    except Exception as _exc:
        print(_exc)
        pass


if __name__ == '__main__':

    sg.theme('DarkTeal2')
    dataframe = None

    layout_fr1 = [[sg.Menu([['Menu', ['Help', 'About', 'Exit']]], key='-MENU-', background_color='lightgrey')],

                  [sg.Text('Choose File', size=(10, 1)),
                   sg.Input(key='-AIR_TUNNEL_INPUT-', size=(75, 1), change_submits=True, visible=True),
                   sg.FileBrowse(
                       file_types=(("CSV Files", "*.csv"),),
                       initial_folder=r'C:/Users/' + os.getlogin() + '/Desktop/',
                       size=(10, 1),
                       key='-AIR_TUNNEL_BROWSE-',
                       tooltip='Only CSV Files')],

                  [sg.Text('Save File', size=(10, 1)),
                   sg.Input(key='-AIR_TUNNEL_SAVE_INPUT-', size=(75, 1), change_submits=True, visible=True),
                   sg.FileSaveAs(
                       file_types=(('Excel File', '*.xlsx'),),
                       initial_folder=r'C:/Users/' + os.getlogin() + '/Desktop/',
                       key='-AIR_TUNNEL_SAVE-',
                       size=(10, 1))]
                  ]
    layout_fr2 = [[sg.Text(text='', size=(95, 1), visible=False)]]
    layout_fr3 = [[sg.Text(text='Choose File', size=(10, 1)),
                   sg.Input(key='-RANGE_ROOM_INPUT-', size=(75, 1), change_submits=True, visible=True),
                   sg.FileBrowse(
                       file_types=(('CSV Files', '*.csv'),),
                       initial_folder=r'C:/Users/' + os.getlogin() + '/Desktop/',
                       size=(10, 1),
                       key='RANGE_ROOM_BROWSE-',
                       tooltip='Only CSV Files')],

                  [sg.Text('Save file', size=(10, 1)),
                   sg.Input(key='-RANGE_ROOM_SAVE_INPUT-', size=(75, 1), change_submits=True, visible=True),
                   sg.FileSaveAs(
                       file_types=(('Excel File', '*.xlsx'),),
                       initial_folder=r'C:/Users/' + os.getlogin() + '/Desktop/',
                       key='-RANGE_ROOM_SAVE-',
                       size=(10, 1))],
                  [sg.FileBrowse(
                      button_text='Process',
                      file_types=(('Excel File', '*.xls*'),),
                      initial_folder=r'C:/Users/' + os.getlogin() + '/Desktop/',
                      key='-PROCESS-', size=(10, 1), change_submits=True)]
                  ]

    frame01 = sg.Frame(title='Air Tunnel', layout=layout_fr1, title_color='orange', font=('Consolas', 12, 'bold'),
                       title_location=sg.TITLE_LOCATION_TOP)
    frame02 = sg.Frame(title='', layout=layout_fr2, relief=sg.RELIEF_FLAT)
    frame03 = sg.Frame(title='Range Room', layout=layout_fr3, title_color='Yellow', font=('Consolas', 12, 'bold'),
                       title_location=sg.TITLE_LOCATION_TOP, element_justification='center')

    # Set from 16:9 to 12:6 or 8:3
    _H, _V = sg.Window.get_screen_size()
    # Make Window
    window = sg.Window(title='Air Tunnel', layout=[[frame01], [frame02], [frame03]], size=(2 * _H // 4, 1 * _V // 3))

    while True:
        event, values = window.read()
        # print(values)
        if event in [sg.WINDOW_CLOSED, 'Exit']:
            break
        if event in '-AIR_TUNNEL_INPUT-':
            if 'Every_Step_Log' in values['-AIR_TUNNEL_INPUT-']:
                dataframe = read_air_tunnel(values['-AIR_TUNNEL_INPUT-'])
                if dataframe.empty:
                    window['-AIR_TUNNEL_INPUT-'].update('')
            else:
                sg.Popup('Select Correct File')
                window['-AIR_TUNNEL_INPUT-'].update('')
                values['-AIR_TUNNEL_INPUT-'] = ''
                values['-AIR_TUNNEL_BROWSE-'] = ''
        if event in '-AIR_TUNNEL_SAVE_INPUT-':
            if values['-AIR_TUNNEL_SAVE_INPUT-']:
                try:
                    save_air_tunnel(dataframe, values['-AIR_TUNNEL_SAVE_INPUT-'])
                except NameError:
                    print(NameError)
                    window['-AIR_TUNNEL_SAVE_INPUT-'].update('')
                    continue
        if event in '-RANGE_ROOM_INPUT-':
            if 'Every_Step_Log' in values['-RANGE_ROOM_INPUT-']:
                range_room_dataframe = read_range_room(values['-RANGE_ROOM_INPUT-'])
                if range_room_dataframe.empty:
                    window['-RANGE_ROOM_INPUT-'].update('')
        if event in '-RANGE_ROOM_SAVE_INPUT-':
            if values['-RANGE_ROOM_SAVE_INPUT-']:
                try:
                    save_range_room(range_room_dataframe, values['-RANGE_ROOM_SAVE_INPUT-'])
                except NameError:
                    print(NameError)
                    window['-RANGE_ROOM_SAVE_INPUT-'].update('')
                    continue
        if event in '-PROCESS-':
            print('Inside Event', process_file(values['-PROCESS-']))
            # sg.Popup('Not Implemented Yet ;)')
            values['-PROCESS-'] = ''
        if event in 'Help':
            sg.Popup('BROWSE : Use to Select & Load csv File.\n'
                     'SAVE AS : Specify Location/Name to Save Charts', no_titlebar=True, background_color='black')
        if event in 'About':
            sg.Popup('© Author : <3nt1t1> \n© Contact : nikoldimitris@gmail.com\n© Author : D. Nikolopoulos \n© Contact : nikoldimitris@gmail.com',
                     no_titlebar=True, background_color='black')
        print(values)
